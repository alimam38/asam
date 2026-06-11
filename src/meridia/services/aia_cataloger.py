#!/usr/bin/env python3
"""
AIA Corpus Cataloger — Layer 1 Automated Ingestion
Eden Intelligence Group

Watches a designated folder on the NAS for new documents.
Extracts text, computes checksums, registers as 'raw' in corpus.documents.
Classification happens in Layer 2 (intelligent sessions with Claude/Gemini).

Deployment: Docker container or cron job on Synology DS925+

Usage:
    # One-time scan of a folder
    python aia_cataloger.py --scan /path/to/dropbox/folder

    # Watch mode (continuous monitoring)
    python aia_cataloger.py --watch /path/to/dropbox/folder

    # Scan with custom database connection
    python aia_cataloger.py --scan /path/to/folder --db-host localhost --db-port 5432 --db-name meridia
"""

import os
import sys
import hashlib
import argparse
import json
import time
import logging
from pathlib import Path
from datetime import datetime, date
from typing import Optional

# ─── Text Extraction ─────────────────────────────────────────────

def extract_text_pdf(filepath: str) -> Optional[str]:
    """Extract text from PDF files."""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(filepath)
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n\n".join(pages) if pages else None
    except Exception as e:
        logging.warning(f"PDF extraction failed for {filepath}: {e}")
        return None


def extract_text_docx(filepath: str) -> Optional[str]:
    """Extract text from Word documents."""
    try:
        from docx import Document
        doc = Document(filepath)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs) if paragraphs else None
    except Exception as e:
        logging.warning(f"DOCX extraction failed for {filepath}: {e}")
        return None


def extract_text_xlsx(filepath: str) -> Optional[str]:
    """Extract text from Excel spreadsheets."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    all_text.append(" | ".join(cells))
        return "\n".join(all_text) if all_text else None
    except Exception as e:
        logging.warning(f"XLSX extraction failed for {filepath}: {e}")
        return None


def extract_text_plain(filepath: str) -> Optional[str]:
    """Extract text from plain text files (md, txt, html, json, sql, py, js, etc.)."""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            return f.read()
    except Exception as e:
        logging.warning(f"Text extraction failed for {filepath}: {e}")
        return None


# Map of file extensions to extraction functions
EXTRACTORS = {
    '.pdf': extract_text_pdf,
    '.docx': extract_text_docx,
    '.doc': extract_text_docx,  # May not work for old .doc format
    '.xlsx': extract_text_xlsx,
    '.xls': extract_text_xlsx,
    '.md': extract_text_plain,
    '.txt': extract_text_plain,
    '.html': extract_text_plain,
    '.htm': extract_text_plain,
    '.json': extract_text_plain,
    '.sql': extract_text_plain,
    '.py': extract_text_plain,
    '.js': extract_text_plain,
    '.css': extract_text_plain,
    '.csv': extract_text_plain,
    '.xml': extract_text_plain,
    '.yaml': extract_text_plain,
    '.yml': extract_text_plain,
    '.toml': extract_text_plain,
    '.ini': extract_text_plain,
    '.cfg': extract_text_plain,
    '.env': extract_text_plain,
    '.sh': extract_text_plain,
    '.bat': extract_text_plain,
}

# Files to skip
SKIP_FILES = {'.ds_store', 'thumbs.db', '.gitignore', '.dropbox', 'desktop.ini'}
SKIP_DIRS = {'.git', 'node_modules', '__pycache__', '.dropbox.cache', '@eaDir'}


# ─── Core Functions ──────────────────────────────────────────────

def compute_checksum(filepath: str) -> str:
    """Compute SHA-256 checksum for deduplication."""
    sha256 = hashlib.sha256()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            sha256.update(chunk)
    return sha256.hexdigest()


def extract_text(filepath: str) -> Optional[str]:
    """Route to appropriate text extractor based on file extension."""
    ext = Path(filepath).suffix.lower()
    extractor = EXTRACTORS.get(ext)
    if extractor:
        return extractor(filepath)
    logging.info(f"No extractor for {ext}, skipping text extraction: {filepath}")
    return None


def infer_created_date(filepath: str) -> Optional[str]:
    """Try to infer original creation date from file metadata."""
    try:
        stat = os.stat(filepath)
        # Use the earlier of creation time and modification time
        ctime = stat.st_ctime
        mtime = stat.st_mtime
        earliest = min(ctime, mtime)
        return date.fromtimestamp(earliest).isoformat()
    except Exception:
        return None


def catalog_file(filepath: str, source_root: str) -> Optional[dict]:
    """
    Process a single file and return a catalog record.
    Returns None if file should be skipped.
    """
    path = Path(filepath)
    
    # Skip hidden files, system files, and directories we don't want
    if path.name.lower() in SKIP_FILES:
        return None
    if path.name.startswith('.'):
        return None
    
    # Get file info
    ext = path.suffix.lower()
    file_size = os.path.getsize(filepath)
    
    # Skip empty files
    if file_size == 0:
        return None
    
    # Compute checksum
    checksum = compute_checksum(filepath)
    
    # Extract text
    full_text = extract_text(filepath)
    
    # Determine file type category
    file_type = ext.lstrip('.') if ext else 'unknown'
    
    # Build relative source path from the root being scanned
    try:
        rel_path = os.path.relpath(filepath, source_root)
    except ValueError:
        rel_path = filepath
    
    # Clean title from filename
    title = path.stem.replace('_', ' ').replace('-', ' ')
    
    # Infer creation date
    created_date = infer_created_date(filepath)
    
    return {
        'title': title,
        'filename': path.name,
        'file_type': file_type,
        'file_size_bytes': file_size,
        'checksum_sha256': checksum,
        'status': 'raw',
        'origin': 'dropbox',  # Default; can be overridden
        'domain': 'general',
        'era': 'pre_meridia',  # Default for historical; classified later
        'confidence': 'uncertain',
        'full_text': full_text,
        'created_date': created_date,
        'source_path': rel_path,
        'nas_path': None,  # Set after files are on NAS
    }


def scan_folder(folder_path: str, origin: str = 'dropbox') -> list[dict]:
    """
    Scan a folder recursively and catalog all supported files.
    Returns list of catalog records.
    """
    records = []
    folder = Path(folder_path)
    
    if not folder.exists():
        logging.error(f"Folder does not exist: {folder_path}")
        return records
    
    for root, dirs, files in os.walk(folder_path):
        # Skip unwanted directories
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]
        
        for filename in sorted(files):
            filepath = os.path.join(root, filename)
            record = catalog_file(filepath, folder_path)
            if record:
                record['origin'] = origin
                records.append(record)
                logging.info(f"Cataloged: {record['filename']} ({record['file_type']}, {record['file_size_bytes']} bytes)")
    
    return records


# ─── Database Operations ─────────────────────────────────────────

def get_db_connection(host='localhost', port=5432, dbname='meridia', user='postgres', password=''):
    """Create PostgreSQL connection."""
    try:
        import psycopg2
        conn = psycopg2.connect(
            host=host,
            port=port,
            dbname=dbname,
            user=user,
            password=password
        )
        return conn
    except ImportError:
        logging.error("psycopg2 not installed. Install with: pip install psycopg2-binary")
        return None
    except Exception as e:
        logging.error(f"Database connection failed: {e}")
        return None


def check_duplicate(conn, checksum: str) -> bool:
    """Check if a document with this checksum already exists."""
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM corpus.documents WHERE checksum_sha256 = %s", (checksum,))
    count = cur.fetchone()[0]
    cur.close()
    return count > 0


def insert_record(conn, record: dict) -> Optional[str]:
    """Insert a catalog record into corpus.documents. Returns document ID."""
    if check_duplicate(conn, record['checksum_sha256']):
        logging.info(f"Duplicate skipped: {record['filename']}")
        return None
    
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO corpus.documents (
            title, filename, file_type, file_size_bytes, checksum_sha256,
            status, origin, domain, era, confidence,
            full_text, created_date, source_path, nas_path
        ) VALUES (
            %(title)s, %(filename)s, %(file_type)s, %(file_size_bytes)s, %(checksum_sha256)s,
            %(status)s, %(origin)s, %(domain)s, %(era)s, %(confidence)s,
            %(full_text)s, %(created_date)s, %(source_path)s, %(nas_path)s
        )
        RETURNING id
    """, record)
    
    doc_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    logging.info(f"Inserted: {record['filename']} → {doc_id}")
    return str(doc_id)


def bulk_insert(conn, records: list[dict]) -> dict:
    """Insert multiple records, skipping duplicates. Returns stats."""
    stats = {'inserted': 0, 'skipped_duplicate': 0, 'failed': 0}
    
    for record in records:
        try:
            result = insert_record(conn, record)
            if result:
                stats['inserted'] += 1
            else:
                stats['skipped_duplicate'] += 1
        except Exception as e:
            logging.error(f"Failed to insert {record['filename']}: {e}")
            conn.rollback()
            stats['failed'] += 1
    
    return stats


def create_session_record(conn, session_type: str, model: str, title: str) -> Optional[str]:
    """Create a session record and return its ID."""
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO corpus.sessions (session_type, model_used, title)
        VALUES (%s, %s, %s)
        RETURNING id
    """, (session_type, model, title))
    session_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    return str(session_id)


# ─── SQL Output Mode (No Database Required) ─────────────────────

def generate_sql_inserts(records: list[dict], output_path: str):
    """
    Generate SQL INSERT statements for records.
    Use this when you can't connect directly to the NAS database.
    Copy the output SQL file to the NAS and run it there.
    """
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("-- AIA Corpus Cataloger — Batch Insert\n")
        f.write(f"-- Generated: {datetime.now().isoformat()}\n")
        f.write(f"-- Records: {len(records)}\n")
        f.write("-- Run this on the NAS PostgreSQL after deploying the corpus schema\n\n")
        f.write("BEGIN;\n\n")
        
        # Create session record
        f.write("-- Session record\n")
        f.write(f"""INSERT INTO corpus.sessions (session_type, model_used, title, documents_added)
VALUES ('ingestion', 'aia_cataloger_v1', 'Batch cataloging - {datetime.now().strftime("%Y-%m-%d")}', {len(records)});\n\n""")
        
        for i, record in enumerate(records):
            # Escape single quotes in text fields
            def esc(val):
                if val is None:
                    return 'NULL'
                if isinstance(val, str):
                    return "'" + val.replace("'", "''") + "'"
                return str(val)
            
            # Truncate full_text for SQL output to keep file manageable
            full_text = record.get('full_text')
            if full_text and len(full_text) > 50000:
                full_text = full_text[:50000] + '\n\n[TRUNCATED — full text available in source file]'
            
            f.write(f"-- [{i+1}/{len(records)}] {record['filename']}\n")
            f.write(f"""INSERT INTO corpus.documents (
    title, filename, file_type, file_size_bytes, checksum_sha256,
    status, origin, domain, era, confidence,
    full_text, created_date, source_path
) VALUES (
    {esc(record['title'])},
    {esc(record['filename'])},
    {esc(record['file_type'])},
    {record['file_size_bytes']},
    {esc(record['checksum_sha256'])},
    {esc(record['status'])},
    {esc(record['origin'])},
    {esc(record['domain'])},
    {esc(record['era'])},
    {esc(record['confidence'])},
    {esc(full_text)},
    {esc(record['created_date'])},
    {esc(record['source_path'])}
) ON CONFLICT DO NOTHING;\n\n""")
        
        f.write("COMMIT;\n")
        f.write(f"\n-- Refresh materialized view\n")
        f.write("SELECT corpus.refresh_canon();\n")
    
    logging.info(f"SQL inserts written to: {output_path}")


# ─── Watch Mode ──────────────────────────────────────────────────

def watch_folder(folder_path: str, db_config: dict, interval: int = 60):
    """
    Continuously watch a folder for new files.
    Checks every `interval` seconds.
    """
    logging.info(f"Watching {folder_path} (checking every {interval}s)")
    known_checksums = set()
    
    # Initial scan to build known set
    conn = get_db_connection(**db_config)
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT checksum_sha256 FROM corpus.documents")
        known_checksums = {row[0] for row in cur.fetchall()}
        cur.close()
        conn.close()
        logging.info(f"Loaded {len(known_checksums)} known checksums from database")
    
    while True:
        try:
            records = scan_folder(folder_path)
            new_records = [r for r in records if r['checksum_sha256'] not in known_checksums]
            
            if new_records:
                logging.info(f"Found {len(new_records)} new files")
                conn = get_db_connection(**db_config)
                if conn:
                    stats = bulk_insert(conn, new_records)
                    logging.info(f"Inserted: {stats['inserted']}, Skipped: {stats['skipped_duplicate']}, Failed: {stats['failed']}")
                    
                    # Update known checksums
                    for r in new_records:
                        known_checksums.add(r['checksum_sha256'])
                    
                    conn.close()
            
            time.sleep(interval)
            
        except KeyboardInterrupt:
            logging.info("Watch stopped by user")
            break
        except Exception as e:
            logging.error(f"Watch cycle error: {e}")
            time.sleep(interval)


# ─── CLI ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='AIA Corpus Cataloger — Automated document ingestion for the NAS'
    )
    
    # Mode
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument('--scan', metavar='FOLDER', help='One-time scan of a folder')
    mode.add_argument('--watch', metavar='FOLDER', help='Continuous watch mode')
    
    # Output
    parser.add_argument('--sql-output', metavar='FILE', 
                       help='Generate SQL INSERT file instead of writing to database directly')
    parser.add_argument('--json-output', metavar='FILE',
                       help='Output catalog records as JSON')
    
    # Database
    parser.add_argument('--db-host', default='localhost')
    parser.add_argument('--db-port', type=int, default=5432)
    parser.add_argument('--db-name', default='meridia')
    parser.add_argument('--db-user', default='postgres')
    parser.add_argument('--db-pass', default='')
    
    # Options
    parser.add_argument('--origin', default='dropbox', 
                       choices=['dropbox', 'claude_session', 'gemini_session', 'gpt_session', 'manual', 'external'],
                       help='Document origin tag')
    parser.add_argument('--interval', type=int, default=60,
                       help='Watch mode check interval in seconds')
    parser.add_argument('--verbose', '-v', action='store_true')
    
    args = parser.parse_args()
    
    # Setup logging
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    db_config = {
        'host': args.db_host,
        'port': args.db_port,
        'dbname': args.db_name,
        'user': args.db_user,
        'password': args.db_pass,
    }
    
    if args.scan:
        # One-time scan
        logging.info(f"Scanning: {args.scan}")
        records = scan_folder(args.scan, origin=args.origin)
        logging.info(f"Found {len(records)} files")
        
        if not records:
            logging.info("No files to catalog")
            return
        
        if args.sql_output:
            generate_sql_inserts(records, args.sql_output)
        elif args.json_output:
            # JSON output for inspection
            json_records = [{k: v for k, v in r.items() if k != 'full_text'} for r in records]
            with open(args.json_output, 'w') as f:
                json.dump(json_records, f, indent=2, default=str)
            logging.info(f"JSON catalog written to: {args.json_output}")
        else:
            # Direct database insert
            conn = get_db_connection(**db_config)
            if conn:
                stats = bulk_insert(conn, records)
                logging.info(f"Complete — Inserted: {stats['inserted']}, Skipped: {stats['skipped_duplicate']}, Failed: {stats['failed']}")
                conn.close()
            else:
                logging.error("Could not connect to database. Use --sql-output to generate SQL file instead.")
    
    elif args.watch:
        watch_folder(args.watch, db_config, interval=args.interval)


if __name__ == '__main__':
    main()
