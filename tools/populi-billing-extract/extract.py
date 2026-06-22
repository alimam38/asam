#!/usr/bin/env python3
"""
Populi billing extract — revenue by account x month (the Populi side of the
Populi -> QuickBooks reconciliation).

Pulls Populi's posted financial transactions for a date range, expands each to
its ledger entries, and aggregates net postings (credit - debit) by income /
contra-revenue account and by month. Writes an .xlsx structured like the QBO
"Turner-FY26-Revenue-Reconciliation" workbook so the two reconcile by account
NAME (Populi and QBO account NUMBERS diverged after the QB restructuring).

Revenue logic (verified against the live API 2026-06-21):
  - Each posted transaction's ledger_entries give account-level debits/credits.
  - Net credit to an INCOME account = revenue (invoices + sales receipts).
  - Customer payments credit A/R (an asset), so they are excluded automatically.
  - Scholarship/contra accounts (booked as liabilities in Populi) are included so
    you see gross vs. aid, mirroring the QBO workbook.

Setup: copy .env.example -> .env and put your key in POPULI_API_KEY (the script
  auto-loads .env). Or pass it inline:
  POPULI_API_KEY=sk_xxx POPULI_SCHOOL=turnerseminary python extract.py
  See README.md — the recommended path is to run this from Claude Code.
Optional env: POPULI_BASE_URL, POPULI_FY_START (default 2025-07-01),
              POPULI_FY_END (default 2026-06-30), OUT (default populi-revenue-by-account.xlsx)

Deps:  pip install -r requirements.txt   (httpx, openpyxl, python-dotenv)
"""
import os, sys, time, datetime as dt
from collections import defaultdict
import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Load a local .env (this script's own folder) if python-dotenv is installed, so the
# Populi key never has to be typed at the prompt. Real environment variables still take
# precedence, and the script still works if python-dotenv isn't present.
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
except ImportError:
    pass

KEY = os.environ.get("POPULI_API_KEY", "").strip()
BASE = os.environ.get("POPULI_BASE_URL", "").strip().rstrip("/")
if not BASE:
    BASE = f"https://{os.environ.get('POPULI_SCHOOL','turnerseminary').strip()}.populiweb.com/api2"
START = os.environ.get("POPULI_FY_START", "2025-07-01")
END = os.environ.get("POPULI_FY_END", "2026-06-30")
OUT = os.environ.get("OUT", "populi-revenue-by-account.xlsx")
HEADERS = {"Authorization": f"Bearer {KEY}", "Accept": "application/json",
           "Content-Type": "application/json"}
MONEY = '$#,##0;($#,##0);-'


def req(method, path, body=None):
    url = f"{BASE}/{path.lstrip('/')}"
    for attempt in range(1, 6):
        r = httpx.request(method, url, headers=HEADERS, json=body, timeout=60)
        if r.status_code == 429:
            time.sleep(min(60, 5 * attempt)); continue
        r.raise_for_status()
        return r.json()
    r.raise_for_status()


def month_label(iso): return dt.date.fromisoformat(iso).strftime("%b %Y")


def months_between(s, e):
    s, e = dt.date.fromisoformat(s), dt.date.fromisoformat(e)
    out, y, m = [], s.year, s.month
    while (y, m) <= (e.year, e.month):
        out.append(dt.date(y, m, 1).strftime("%b %Y"))
        m += 1
        if m > 12: m, y = 1, y + 1
    return out


def main():
    if not KEY:
        sys.exit("Set POPULI_API_KEY (and POPULI_SCHOOL or POPULI_BASE_URL).")
    print(f"Populi billing extract  {START} -> {END}  @ {BASE}")

    # 1) chart of accounts -> id: {name, number, type}
    accts, page = {}, 1
    while True:
        d = req("GET", "accounts", {"page": page, "limit": 200})
        for a in d["data"]:
            accts[a["id"]] = {"name": a["name"], "number": a.get("account_number"),
                              "type": a.get("type")}
        if not d.get("has_more"): break
        page += 1
    print(f"  {len(accts)} accounts")

    # 2) posted transactions in range
    txids, page = [], 1
    while True:
        d = req("GET", "transactions", {"posted_date_start": START, "posted_date_end": END,
                                        "page": page, "limit": 200})
        for t in d["data"]:
            if t.get("status") == "posted" and not t.get("voided_at"):
                txids.append((t["id"], t["posted_on"]))
        if not d.get("has_more"): break
        page += 1
    print(f"  {len(txids)} posted transactions")

    # 3) expand ledger entries; aggregate net (credit - debit) by account x month
    agg = defaultdict(float)
    for i, (tid, posted) in enumerate(txids, 1):
        d = req("GET", f"transactions/{tid}", {"expand": ["ledger_entries"]})
        mk = month_label(posted)
        for le in d.get("ledger_entries", []):
            agg[(le["account_id"], mk)] += (le.get("credit") or 0) - (le.get("debit") or 0)
        if i % 25 == 0: print(f"    expanded {i}/{len(txids)}")

    # 4) revenue-relevant accounts: income + scholarship/contra
    months = months_between(START, END)
    rev_ids = sorted(
        {aid for (aid, _) in agg
         if accts.get(aid, {}).get("type") == "income"
         or "scholarship" in (accts.get(aid, {}).get("name", "").lower())},
        key=lambda a: (accts.get(a, {}).get("number") or "", accts.get(a, {}).get("name") or ""))

    # 5) write workbook
    wb = Workbook(); ws = wb.active; ws.title = "Populi Revenue by Account"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"Populi revenue by account x month  ({START} to {END})"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="4B2E83")
    ws["A2"] = ("Net credit to income/contra accounts from posted ledger entries. "
                "Reconcile to QuickBooks by account NAME (numbers diverge). Housing posts to Populi "
                "acct 40005-01 'Auxiliary Enterprise Revenue' = QBO 41064-01 Housing Rentals.")
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="6B6878")
    hdr = 4
    headers = ["Account (Populi)", "Populi #"] + months + ["FY Total"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hdr, c, h); cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="4B2E83")
    r = hdr + 1
    first_data = r
    for aid in rev_ids:
        info = accts.get(aid, {})
        ws.cell(r, 1, info.get("name")); ws.cell(r, 2, info.get("number"))
        for j, m in enumerate(months):
            cell = ws.cell(r, 3 + j, round(agg.get((aid, m), 0.0), 2))
            cell.font = Font(name="Arial", size=10, color="0000FF"); cell.number_format = MONEY
        last_m = get_column_letter(2 + len(months))
        fc = ws.cell(r, 3 + len(months), f"=SUM(C{r}:{last_m}{r})")
        fc.font = Font(name="Arial", size=10, bold=True); fc.number_format = MONEY
        r += 1
    # total row
    ws.cell(r, 1, "TOTAL REVENUE (Populi)").font = Font(name="Arial", bold=True, color="FFFFFF")
    for c in range(1, 3 + len(months) + 1): ws.cell(r, c).fill = PatternFill("solid", start_color="4B2E83")
    for c in range(3, 3 + len(months) + 1):
        L = get_column_letter(c)
        tc = ws.cell(r, c, f"=SUM({L}{first_data}:{L}{r-1})")
        tc.font = Font(name="Arial", bold=True, color="FFFFFF"); tc.number_format = MONEY
    ws.column_dimensions["A"].width = 48; ws.column_dimensions["B"].width = 16
    for c in range(3, 3 + len(months) + 1): ws.column_dimensions[get_column_letter(c)].width = 12
    ws.freeze_panes = "C5"
    wb.save(OUT)
    print(f"Saved {OUT}  ({len(rev_ids)} revenue accounts x {len(months)} months)")
    print("Next: diff this against the QBO 'Turner-FY26-Revenue-Reconciliation' workbook by account NAME.")


if __name__ == "__main__":
    main()
